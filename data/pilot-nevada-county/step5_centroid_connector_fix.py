#!/usr/bin/env python3
"""Step 5: Repair Grass Valley centroid connectors and rerun assignment/validation.

This script resets the AequilibraE project DB to a clean pre-step5 seed, rewires the
problem Grass Valley-area centroid connectors onto consistent highway-feeder nodes,
reruns the v2 demand assignment, and writes refreshed outputs to run_output_v3/.
"""

from __future__ import annotations

import json
import os
import shutil
import sqlite3
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

os.environ["SPATIALITE_LIBRARY_PATH"] = "/home/linuxbrew/.linuxbrew/lib/mod_spatialite"

from aequilibrae import Project
from aequilibrae.matrix import AequilibraeMatrix
from aequilibrae.paths import NetworkSkimming, TrafficAssignment, TrafficClass


DATA_DIR = Path(__file__).resolve().parent
PROJ_DIR = DATA_DIR / "aeq_project"
DB_PATH = PROJ_DIR / "project_database.sqlite"
DB_SEED_PATH = PROJ_DIR / "project_database_step5_seed.sqlite"
PKG_DIR = DATA_DIR / "package"
V2_OUT = DATA_DIR / "run_output_v2"
OUT_DIR = DATA_DIR / "run_output_v3"
CACHE_PATH = DATA_DIR / "caltrans_2023_aadt.xlsx"

OUT_DIR.mkdir(exist_ok=True)

# Manually curated, geometry-consistent highway feeder nodes around Grass Valley.
# Each node was checked to ensure its incident-link geometry matches the node
# geometry, which avoids the corrupted mixed-node attachments left by step 2.
TARGET_ZONE_FIXES: dict[int, list[int]] = {
    # Grass Valley core / SR 20 at Jct 49 area
    1: [5127, 5112, 9612],
    4: [9612, 5112, 5127],
    # South-central Grass Valley / SR 174
    8: [3075, 9804, 7508],
    # North Grass Valley / Nevada City Hwy / SR 49 north
    19: [10794, 5966, 3206],
}

STATIONS = [
    {
        "station_id": "CT_SR20_JCT49",
        "label": "SR 20 at Jct Rte 49",
        "caltrans": {"rte": "020", "cnty": "NEV", "desc_contains": "JCT. RTE. 49"},
        "osm_names": ["Golden Center Freeway", "Golden Center Expressway", "Grass Valley Highway"],
        "bbox": (-121.06, 39.215, -121.04, 39.235),
    },
    {
        "station_id": "CT_SR20_BRUNSWICK",
        "label": "SR 20 at Brunswick Rd",
        "caltrans": {"rte": "020", "cnty": "NEV", "desc_contains": "BRUNSWICK ROAD"},
        "osm_names": ["Grass Valley Highway", "Brunswick Road"],
        "bbox": (-121.05, 39.230, -121.03, 39.245),
    },
    {
        "station_id": "CT_SR20_PENNVALLEY",
        "label": "SR 20 at Penn Valley Dr",
        "caltrans": {"rte": "020", "cnty": "NEV", "desc_contains": "PENN VALLEY DRIVE"},
        "osm_names": ["Penn Valley Drive", "Grass Valley Highway"],
        "bbox": (-121.19, 39.190, -121.16, 39.210),
    },
    {
        "station_id": "CT_SR49_SOUTHGV",
        "label": "SR 49 at South Grass Valley",
        "caltrans": {"rte": "049", "cnty": "NEV", "desc_contains": "SOUTH GRASS VALLEY"},
        "osm_names": ["State Highway 49", "Alta Sierra Drive", "Golden Center Freeway"],
        "bbox": (-121.08, 39.130, -121.04, 39.170),
    },
    {
        "station_id": "CT_SR174_BRUNSWICK",
        "label": "SR 174 at Brunswick Rd",
        "caltrans": {"rte": "174", "cnty": "NEV", "desc_contains": "BRUNSWICK ROAD"},
        "osm_names": ["Colfax Highway", "Brunswick Road"],
        "bbox": (-121.06, 39.200, -121.03, 39.225),
    },
]


def spatial_conn(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.enable_load_extension(True)
    conn.load_extension("/home/linuxbrew/.linuxbrew/lib/mod_spatialite")
    return conn


def ensure_seed_copy() -> None:
    if not DB_SEED_PATH.exists():
        shutil.copy2(DB_PATH, DB_SEED_PATH)
    shutil.copy2(DB_SEED_PATH, DB_PATH)


def zone_to_centroid_map(conn: sqlite3.Connection) -> dict[int, int]:
    zones = pd.read_csv(PKG_DIR / "zone_attributes.csv")
    mapping: dict[int, int] = {}
    centroids = conn.execute(
        "SELECT node_id, X(geometry), Y(geometry) FROM nodes WHERE is_centroid=1 ORDER BY node_id"
    ).fetchall()
    for node_id, x, y in centroids:
        match = zones[(zones["centroid_lon"] == x) & (zones["centroid_lat"] == y)]
        if len(match) != 1:
            raise RuntimeError(f"Could not uniquely match centroid node {node_id}")
        mapping[int(match.iloc[0]["zone_id"])] = int(node_id)
    return mapping


def node_consistency_ratio(conn: sqlite3.Connection, node_id: int) -> float:
    row = conn.execute("SELECT X(geometry), Y(geometry) FROM nodes WHERE node_id=?", (node_id,)).fetchone()
    if row is None:
        raise RuntimeError(f"Node {node_id} not found")
    x, y = row
    links = conn.execute(
        """
        SELECT link_id, a_node, b_node,
               X(StartPoint(geometry)), Y(StartPoint(geometry)),
               X(EndPoint(geometry)), Y(EndPoint(geometry))
        FROM links
        WHERE a_node=? OR b_node=?
        """,
        (node_id, node_id),
    ).fetchall()
    if not links:
        return 0.0
    ok = 0
    for _, a_node, b_node, sx, sy, ex, ey in links:
        matched = (a_node == node_id and abs(sx - x) < 1e-5 and abs(sy - y) < 1e-5) or (
            b_node == node_id and abs(ex - x) < 1e-5 and abs(ey - y) < 1e-5
        )
        ok += int(matched)
    return ok / len(links)


def node_major_names(conn: sqlite3.Connection, node_id: int) -> list[str]:
    rows = conn.execute(
        """
        SELECT DISTINCT COALESCE(name, ''), link_type
        FROM links
        WHERE (a_node=? OR b_node=?)
          AND link_type IN ('motorway', 'trunk', 'primary', 'secondary')
        ORDER BY link_type, name
        """,
        (node_id, node_id),
    ).fetchall()
    names = []
    for name, link_type in rows:
        label = name if name else f"<{link_type}>"
        names.append(label)
    return names


def ensure_connector_link_type(conn: sqlite3.Connection) -> None:
    existing = conn.execute(
        "SELECT 1 FROM link_types WHERE link_type='centroid_connector'"
    ).fetchone()
    if existing:
        return
    conn.execute(
        """
        INSERT INTO link_types (link_type, link_type_id, description, lanes, lane_capacity)
        VALUES ('centroid_connector', 'z', 'Virtual centroid connectors', 10, 10000)
        """
    )
    conn.commit()


def remove_existing_connectors(conn: sqlite3.Connection, centroid_id: int) -> int:
    count = conn.execute(
        "SELECT COUNT(*) FROM links WHERE link_type='centroid_connector' AND (a_node=? OR b_node=?)",
        (centroid_id, centroid_id),
    ).fetchone()[0]
    conn.execute(
        "DELETE FROM links WHERE link_type='centroid_connector' AND (a_node=? OR b_node=?)",
        (centroid_id, centroid_id),
    )
    return int(count)


def haversine_m(lon1: float, lat1: float, lon2: float, lat2: float) -> float:
    r = 6371000.0
    p1 = np.radians(lat1)
    p2 = np.radians(lat2)
    dp = np.radians(lat2 - lat1)
    dl = np.radians(lon2 - lon1)
    a = np.sin(dp / 2) ** 2 + np.cos(p1) * np.cos(p2) * np.sin(dl / 2) ** 2
    return float(2 * r * np.arctan2(np.sqrt(a), np.sqrt(1 - a)))


def add_connector(conn: sqlite3.Connection, link_id: int, centroid_id: int, node_id: int) -> dict:
    cx, cy = conn.execute(
        "SELECT X(geometry), Y(geometry) FROM nodes WHERE node_id=?", (centroid_id,)
    ).fetchone()
    nx, ny = conn.execute(
        "SELECT X(geometry), Y(geometry) FROM nodes WHERE node_id=?", (node_id,)
    ).fetchone()
    line_wkt = f"LINESTRING({cx} {cy}, {nx} {ny})"
    length_m = max(haversine_m(cx, cy, nx, ny), 10.0)
    speed_kph = 50.0
    travel_time_min = max(length_m / (speed_kph * 1000.0 / 60.0), 0.01)
    conn.execute(
        """
        INSERT INTO links (
            link_id, a_node, b_node, direction, distance, modes,
            link_type, name, speed_ab, speed_ba, travel_time_ab, travel_time_ba,
            capacity_ab, capacity_ba, lanes_ab, lanes_ba, geometry
        )
        VALUES (?, ?, ?, 0, ?, 'c', 'centroid_connector', ?, ?, ?, ?, ?, 99999, 99999, 1, 1, GeomFromText(?, 4326))
        """,
        (
            link_id,
            centroid_id,
            node_id,
            length_m,
            f"z{centroid_id}_to_{node_id}",
            speed_kph,
            speed_kph,
            travel_time_min,
            travel_time_min,
            line_wkt,
        ),
    )
    return {
        "link_id": int(link_id),
        "to_node": int(node_id),
        "distance_m": round(length_m, 1),
        "node_major_names": node_major_names(conn, node_id),
        "node_consistency_ratio": round(node_consistency_ratio(conn, node_id), 3),
    }


def apply_connector_fixes() -> dict:
    ensure_seed_copy()
    conn = spatial_conn(DB_PATH)
    ensure_connector_link_type(conn)
    z2c = zone_to_centroid_map(conn)
    next_link = int(conn.execute("SELECT MAX(link_id) FROM links").fetchone()[0]) + 1

    fix_report = {"zones": []}
    print("\n[1] Rewiring Grass Valley centroid connectors...")
    for zone_id, target_nodes in TARGET_ZONE_FIXES.items():
        centroid_id = z2c[zone_id]
        removed = remove_existing_connectors(conn, centroid_id)
        added = []
        for node_id in target_nodes:
            ratio = node_consistency_ratio(conn, node_id)
            if ratio < 0.999:
                raise RuntimeError(
                    f"Refusing to attach zone {zone_id} to inconsistent node {node_id} (ratio={ratio:.3f})"
                )
            added.append(add_connector(conn, next_link, centroid_id, node_id))
            next_link += 1
        fix_report["zones"].append(
            {
                "zone_id": int(zone_id),
                "centroid_id": int(centroid_id),
                "removed_connector_count": int(removed),
                "added_connectors": added,
            }
        )
        summary = ", ".join(str(x["to_node"]) for x in added)
        print(f"  Zone {zone_id:>2d} centroid {centroid_id}: removed {removed}, added -> {summary}")

    conn.commit()
    conn.close()

    report_path = OUT_DIR / "connector_rewire.json"
    with open(report_path, "w") as f:
        json.dump(fix_report, f, indent=2)
    print(f"  Wrote {report_path}")
    return fix_report


def load_v2_demand() -> pd.DataFrame:
    od = pd.read_csv(V2_OUT / "od_trip_matrix_v2.csv", index_col=0)
    od.index = od.index.astype(int)
    od.columns = od.columns.astype(int)
    return od.sort_index().sort_index(axis=1)


def build_assignment_matrix(z2c: dict[int, int]) -> tuple[np.ndarray, list[int], list[int]]:
    od = load_v2_demand()
    zone_ids = sorted(z2c)
    centroids = [z2c[z] for z in zone_ids]
    matrix = od.loc[zone_ids, zone_ids].to_numpy(dtype=float)
    return matrix, centroids, zone_ids


def run_assignment_v3(z2c: dict[int, int]) -> tuple[pd.DataFrame, dict]:
    print("\n[2] Running assignment with repaired connectors...")
    od_array, centroids, zone_ids = build_assignment_matrix(z2c)

    project = Project()
    project.open(str(PROJ_DIR))
    project.network.build_graphs(modes=["c"])
    graph = project.network.graphs["c"]

    cost_field = "travel_time" if "travel_time" in graph.graph.columns else "distance"
    graph.set_graph(cost_field)
    graph.prepare_graph(np.array(centroids))
    graph.set_blocked_centroid_flows(True)
    graph.set_skimming([cost_field])

    demand_path = str(OUT_DIR / "demand_v3.omx")
    if Path(demand_path).exists():
        Path(demand_path).unlink()
    demand_mat = AequilibraeMatrix()
    demand_mat.create_empty(
        file_name=demand_path,
        zones=len(centroids),
        matrix_names=["demand"],
        memory_only=False,
    )
    demand_mat.index = np.array(centroids)
    peak_hour_factor = 0.10
    demand_mat.matrix["demand"][:, :] = od_array * peak_hour_factor
    demand_mat.computational_view(["demand"])

    assig = TrafficAssignment()
    tc = TrafficClass(name="car", graph=graph, matrix=demand_mat)
    tc.set_pce(1.0)
    assig.add_class(tc)
    assig.set_vdf("BPR")
    assig.set_vdf_parameters({"alpha": 0.15, "beta": 4.0})
    assig.set_capacity_field("capacity")
    assig.set_time_field(cost_field)
    assig.max_iter = 50
    assig.rgap_target = 0.01
    assig.set_algorithm("bfw")
    assig.execute()

    rgap = getattr(assig.assignment, "rgap", float("nan"))
    iters = getattr(assig.assignment, "iteration", 50)
    print(f"  Graph: {graph.num_nodes} nodes, {graph.num_links} links, cost={cost_field}")
    print(f"  Demand: {od_array.sum():,.0f} daily trips (assigned at {peak_hour_factor:.0%} peak-hour factor)")
    print(f"  Assignment gap={rgap:.6f}, iterations={iters}")

    results = assig.results()
    lr = results.get_load_results() if hasattr(results, "get_load_results") else results
    if "link_id" not in lr.columns and lr.index.name == "link_id":
        lr = lr.reset_index()
    elif "link_id" not in lr.columns:
        lr = lr.reset_index().rename(columns={"index": "link_id"})

    if "PCE_tot" in lr.columns:
        lr["PCE_tot"] = lr["PCE_tot"] / peak_hour_factor

    skimming = NetworkSkimming(graph)
    skimming.execute()
    skim_mat = skimming.results.skims
    skim_mat.export(str(OUT_DIR / "travel_time_skims_v3.omx"))

    loaded = lr[lr["PCE_tot"] > 0].copy()
    print(f"  Loaded links: {len(loaded)}/{len(lr)}")
    if not loaded.empty:
        print("  Top 10 loaded links:")
        top10 = loaded.nlargest(10, "PCE_tot")
        for _, row in top10.iterrows():
            print(f"    link {int(row['link_id'])}: {row['PCE_tot']:,.0f}")

    lr.to_csv(OUT_DIR / "link_volumes.csv", index=False)
    build_top_loaded_geojson(loaded)

    evidence = {
        "run_id": "nevada-county-pilot-v3-centroid-fix",
        "engine": "AequilibraE 1.6.1",
        "network_source": "OpenStreetMap",
        "model_area": "Nevada County (-121.30,39.00 to -120.00,39.50)",
        "algorithm": "BFW",
        "vdf": "BPR (α=0.15, β=4.0)",
        "convergence": {
            "final_gap": float(rgap) if np.isfinite(rgap) else None,
            "iterations": int(iters),
        },
        "network": {
            "zones": len(zone_ids),
            "centroids": len(centroids),
            "links": int(graph.num_links),
            "nodes": int(graph.num_nodes),
        },
        "demand": {
            "total_trips": round(float(od_array.sum())),
            "source": "run_output_v2/od_trip_matrix_v2.csv",
            "peak_hour_factor": peak_hour_factor,
        },
        "caveats": [
            "Uncalibrated demand (same v2 demand matrix)",
            "OSM default speeds/capacities",
            "Connector repair focused on Grass Valley-area centroid attachment defects",
            "Screening-grade",
        ],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    with open(OUT_DIR / "evidence_packet.json", "w") as f:
        json.dump(evidence, f, indent=2)

    project.close()
    return lr, evidence


def build_top_loaded_geojson(loaded: pd.DataFrame) -> None:
    conn = spatial_conn(DB_PATH)
    features = []
    for _, row in loaded.nlargest(20, "PCE_tot").iterrows():
        lid = int(row["link_id"])
        hit = conn.execute(
            "SELECT name, link_type, AsGeoJSON(geometry) FROM links WHERE link_id=?",
            (lid,),
        ).fetchone()
        if hit is None:
            continue
        name, link_type, geojson = hit
        features.append(
            {
                "type": "Feature",
                "properties": {
                    "link_id": lid,
                    "name": name,
                    "link_type": link_type,
                    "pce_tot": round(float(row["PCE_tot"])),
                },
                "geometry": json.loads(geojson),
            }
        )
    conn.close()
    with open(OUT_DIR / "top_loaded_links.geojson", "w") as f:
        json.dump({"type": "FeatureCollection", "features": features}, f)


def load_caltrans_nev_df() -> pd.DataFrame:
    wb_bytes = CACHE_PATH.read_bytes()
    wb = openpyxl.load_workbook(BytesIO(wb_bytes), read_only=True, data_only=True)
    ws = wb["2023 AADT DATA"]
    header = [
        str(c.value).strip() if c.value else f"col_{i}"
        for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
    ]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = dict(zip(header, row))
        if str(vals.get("CNTY", "")).strip().upper() == "NEV":
            records.append(vals)
    wb.close()
    df = pd.DataFrame(records)
    df["RTE"] = df["RTE"].astype(str).str.strip().str.zfill(3)
    return df


def _safe_int(value) -> int | None:
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", " "):
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def find_best_model_link(
    conn: sqlite3.Connection,
    link_volumes: pd.DataFrame,
    osm_names: list[str],
    bbox: tuple[float, float, float, float],
) -> dict | None:
    xmin, ymin, xmax, ymax = bbox
    name_clauses = " OR ".join(["name = ?" for _ in osm_names])
    sql = f"""
        SELECT link_id, name, link_type,
               X(Centroid(geometry)) AS cx, Y(Centroid(geometry)) AS cy
        FROM links
        WHERE ({name_clauses})
          AND X(Centroid(geometry)) BETWEEN ? AND ?
          AND Y(Centroid(geometry)) BETWEEN ? AND ?
    """
    params = list(osm_names) + [xmin, xmax, ymin, ymax]
    rows = conn.execute(sql, params).fetchall()
    if not rows:
        return None

    best = None
    best_vol = -1.0
    for lid, name, link_type, cx, cy in rows:
        hit = link_volumes.loc[link_volumes["link_id"] == lid, "PCE_tot"]
        vol = float(hit.iloc[0]) if len(hit) else 0.0
        if vol > best_vol:
            best_vol = vol
            best = {
                "link_id": int(lid),
                "name": name,
                "link_type": link_type,
                "pce_tot": round(vol),
                "lon": float(cx),
                "lat": float(cy),
            }
    return best


def run_validation(link_volumes: pd.DataFrame, evidence: dict) -> tuple[pd.DataFrame, dict]:
    print("\n[3] Validating against Caltrans 2023 AADT...")
    nev_df = load_caltrans_nev_df()
    conn = spatial_conn(DB_PATH)

    results = []
    for station in STATIONS:
        rte = station["caltrans"]["rte"]
        desc_sub = station["caltrans"]["desc_contains"].upper()
        ct_match = None
        for _, row in nev_df[nev_df["RTE"] == rte].iterrows():
            desc = str(row.get("DESCRIPTION", "")).strip().upper()
            if desc_sub in desc:
                back = _safe_int(row.get("BACK_AADT"))
                ahead = _safe_int(row.get("AHEAD_AADT"))
                vals = [v for v in (back, ahead) if v is not None]
                if vals:
                    ct_match = {
                        "description": str(row.get("DESCRIPTION", "")).strip(),
                        "observed_aadt": max(vals),
                        "postmile": row.get("PM"),
                    }
                break
        if ct_match is None:
            results.append({
                "station_id": station["station_id"],
                "label": station["label"],
                "match_status": "caltrans_miss",
            })
            continue

        model_link = find_best_model_link(conn, link_volumes, station["osm_names"], station["bbox"])
        if model_link is None:
            results.append({
                "station_id": station["station_id"],
                "label": station["label"],
                "match_status": "model_miss",
                "observed_aadt": ct_match["observed_aadt"],
            })
            continue

        modeled = int(model_link["pce_tot"])
        observed = int(ct_match["observed_aadt"])
        abs_diff = abs(modeled - observed)
        pct_err = 100.0 * abs_diff / observed if observed > 0 else None
        ratio = modeled / observed if observed > 0 else None
        results.append(
            {
                "station_id": station["station_id"],
                "label": station["label"],
                "match_status": "matched",
                "caltrans_description": ct_match["description"],
                "caltrans_postmile": ct_match["postmile"],
                "observed_aadt": observed,
                "model_link_id": model_link["link_id"],
                "model_link_name": model_link["name"],
                "model_link_type": model_link["link_type"],
                "model_lon": round(model_link["lon"], 5),
                "model_lat": round(model_link["lat"], 5),
                "modeled_daily_pce": modeled,
                "absolute_difference": round(abs_diff),
                "absolute_percent_error": round(pct_err, 2) if pct_err is not None else None,
                "volume_ratio_model_obs": round(ratio, 4) if ratio is not None else None,
            }
        )
        ape_txt = f"{pct_err:.1f}%" if pct_err is not None else "N/A"
        print(
            f"  {station['label']}: obs={observed:,} | model={modeled:,} | APE={ape_txt} | link={model_link['link_id']} {model_link['name']}"
        )

    conn.close()

    results_df = pd.DataFrame(results)
    matched = results_df[results_df["match_status"] == "matched"].copy()
    if matched.empty:
        raise RuntimeError("No matched validation stations in v3 run")

    apes = matched["absolute_percent_error"].dropna()
    median_ape = float(apes.median())
    mean_ape = float(apes.mean())
    min_ape = float(apes.min())
    max_ape = float(apes.max())

    matched_sorted = matched.sort_values("observed_aadt", ascending=False)
    obs_rank = matched_sorted["observed_aadt"].rank(ascending=False, method="min")
    mod_rank = matched_sorted["modeled_daily_pce"].rank(ascending=False, method="min")
    n = len(obs_rank)
    if n > 1:
        d_sq = ((obs_rank.values - mod_rank.values) ** 2).sum()
        spearman_rho = 1.0 - (6.0 * d_sq) / (n * (n**2 - 1))
    else:
        spearman_rho = None

    ranking_table = []
    for _, row in matched_sorted.iterrows():
        ranking_table.append(
            {
                "station": row["label"],
                "observed_aadt": int(row["observed_aadt"]),
                "modeled_daily_pce": int(row["modeled_daily_pce"]),
                "obs_rank": int(obs_rank.loc[row.name]),
                "mod_rank": int(mod_rank.loc[row.name]),
            }
        )

    results_df.to_csv(OUT_DIR / "validation_results.csv", index=False)

    v2_summary = json.loads((V2_OUT / "validation_summary.json").read_text())
    summary = {
        "validation_type": "screening_assignment_vs_caltrans_aadt",
        "demand_version": "v3_centroid_connector_fix",
        "caltrans_year": 2023,
        "model_run_id": evidence["run_id"],
        "model_engine": evidence["engine"],
        "model_caveats": evidence["caveats"],
        "stations_total": len(STATIONS),
        "stations_matched": int(len(matched)),
        "stations_missed": int(len(results_df) - len(matched)),
        "metrics": {
            "median_absolute_percent_error": round(median_ape, 2),
            "mean_absolute_percent_error": round(mean_ape, 2),
            "min_absolute_percent_error": round(min_ape, 2),
            "max_absolute_percent_error": round(max_ape, 2),
            "spearman_rho_facility_ranking": round(spearman_rho, 4) if spearman_rho is not None else None,
        },
        "facility_ranking": ranking_table,
        "comparison_to_v2": {
            "v2_median_ape": v2_summary["metrics"]["median_absolute_percent_error"],
            "v2_mean_ape": v2_summary["metrics"]["mean_absolute_percent_error"],
            "v2_spearman_rho": v2_summary["metrics"]["spearman_rho_facility_ranking"],
            "v3_median_ape": round(median_ape, 2),
            "v3_mean_ape": round(mean_ape, 2),
            "v3_spearman_rho": round(spearman_rho, 4) if spearman_rho is not None else None,
        },
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    with open(OUT_DIR / "validation_summary.json", "w") as f:
        json.dump(summary, f, indent=2)

    return results_df, summary


def main() -> int:
    print("=" * 68)
    print("STEP 5: GRASS VALLEY CENTROID CONNECTOR FIX")
    print("=" * 68)

    if not CACHE_PATH.exists():
        print(f"Missing cached workbook: {CACHE_PATH}")
        return 1
    if not (V2_OUT / "od_trip_matrix_v2.csv").exists():
        print(f"Missing v2 demand input: {V2_OUT / 'od_trip_matrix_v2.csv'}")
        return 1

    fix_report = apply_connector_fixes()
    conn = spatial_conn(DB_PATH)
    z2c = zone_to_centroid_map(conn)
    conn.close()

    link_volumes, evidence = run_assignment_v3(z2c)
    validation_results, validation_summary = run_validation(link_volumes, evidence)

    sr20_jct49 = validation_results.loc[
        validation_results["station_id"] == "CT_SR20_JCT49", "modeled_daily_pce"
    ]
    sr20_jct49_volume = int(sr20_jct49.iloc[0]) if len(sr20_jct49) else 0
    median_ape = validation_summary["metrics"]["median_absolute_percent_error"]

    print("\n" + "=" * 68)
    print("STEP 5 COMPLETE")
    print("=" * 68)
    print(f"Fixed zones: {[z['zone_id'] for z in fix_report['zones']]}")
    print(f"Median APE: {median_ape:.2f}%")
    print(f"SR 20 at Jct Rte 49 modeled volume: {sr20_jct49_volume:,}")
    print(f"Below 60% APE: {'YES' if median_ape < 60 else 'NO'}")
    print(f"Output directory: {OUT_DIR}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
