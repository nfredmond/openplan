#!/usr/bin/env python3
"""
Step 3: Observed-count validation against Caltrans 2023 AADT.

Downloads the Caltrans 2023 AADT workbook, extracts 5 priority Nevada County
count stations, matches them to the canonical AequilibraE run output by
geographic proximity + name crosswalk, and produces:
  - run_output/validation_results.csv
  - run_output/validation_summary.json

Requires: step1_osm.py and step2_assign.py to have been run first.
"""

import json
import os
import sqlite3
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import requests

# ── paths ────────────────────────────────────────────────────────────
DATA_DIR = Path(__file__).resolve().parent
OUT_DIR = DATA_DIR / "run_output"
DB_PATH = DATA_DIR / "aeq_project" / "project_database.sqlite"

CALTRANS_URL = (
    "https://dot.ca.gov/-/media/dot-media/programs/traffic-operations/"
    "documents/census/2023/2023-traffic-volumes-ca-a11y.xlsx"
)
CACHE_PATH = DATA_DIR / "caltrans_2023_aadt.xlsx"

# ── 5 priority count stations ───────────────────────────────────────
# Each entry:
#   caltrans_key : (route, county, description_substring)
#   plus the OSM link-name and bounding box used for the spatial crosswalk.
#
# The "observed AADT" for each station is max(BACK_AADT, AHEAD_AADT) from the
# workbook, which gives the higher-side two-way daily volume at the count point
# (conservative for a screening comparison).

STATIONS = [
    {
        "station_id": "CT_SR20_JCT49",
        "label": "SR 20 at Jct Rte 49",
        "caltrans": {"rte": "020", "cnty": "NEV", "desc_contains": "JCT. RTE. 49"},
        "osm_names": ["Golden Center Freeway", "Golden Center Expressway",
                       "Grass Valley Highway"],
        "bbox": (-121.06, 39.215, -121.04, 39.235),
    },
    {
        "station_id": "CT_SR20_BRUNSWICK",
        "label": "SR 20 at Brunswick Rd",
        "caltrans": {"rte": "020", "cnty": "NEV", "desc_contains": "BRUNSWICK ROAD"},
        "osm_names": ["Grass Valley Highway", "Brunswick Road"],
        "bbox": (-121.05, 39.230, -121.03, 39.245),
    },
    {
        "station_id": "CT_SR20_PENNVALLEY",
        "label": "SR 20 at Penn Valley Dr",
        "caltrans": {"rte": "020", "cnty": "NEV", "desc_contains": "PENN VALLEY DRIVE"},
        "osm_names": ["Penn Valley Drive", "Grass Valley Highway"],
        "bbox": (-121.19, 39.190, -121.16, 39.210),
    },
    {
        "station_id": "CT_SR49_SOUTHGV",
        "label": "SR 49 at South Grass Valley",
        "caltrans": {"rte": "049", "cnty": "NEV", "desc_contains": "SOUTH GRASS VALLEY"},
        "osm_names": ["State Highway 49", "Alta Sierra Drive",
                       "Golden Center Freeway"],
        "bbox": (-121.08, 39.130, -121.04, 39.170),
    },
    {
        "station_id": "CT_SR174_BRUNSWICK",
        "label": "SR 174 at Brunswick Rd",
        "caltrans": {"rte": "174", "cnty": "NEV", "desc_contains": "BRUNSWICK ROAD"},
        "osm_names": ["Colfax Highway", "Brunswick Road"],
        "bbox": (-121.06, 39.200, -121.03, 39.225),
    },
]


# ── helpers ──────────────────────────────────────────────────────────

def download_workbook() -> bytes:
    """Download (or read cached) the Caltrans AADT workbook."""
    if CACHE_PATH.exists():
        print(f"  Using cached workbook: {CACHE_PATH}")
        return CACHE_PATH.read_bytes()
    print(f"  Downloading workbook from Caltrans...")
    resp = requests.get(CALTRANS_URL, timeout=120)
    resp.raise_for_status()
    CACHE_PATH.write_bytes(resp.content)
    print(f"  Saved {len(resp.content):,} bytes → {CACHE_PATH}")
    return resp.content


def extract_nev_rows(wb_bytes: bytes) -> pd.DataFrame:
    """Return a DataFrame of all NEV-county rows from the AADT sheet."""
    wb = openpyxl.load_workbook(BytesIO(wb_bytes), read_only=True, data_only=True)
    ws = wb["2023 AADT DATA"]

    # Header is row 1
    header = [str(c.value).strip() if c.value else f"col_{i}"
              for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = dict(zip(header, row))
        if str(vals.get("CNTY", "")).strip().upper() == "NEV":
            records.append(vals)

    wb.close()
    df = pd.DataFrame(records)
    # Normalise route to 3-char zero-padded string
    df["RTE"] = df["RTE"].astype(str).str.strip().str.zfill(3)
    return df


def _safe_int(v) -> int | None:
    """Parse a workbook cell to int, returning None if blank/non-numeric."""
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s == " ":
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def match_caltrans_row(nev_df: pd.DataFrame, spec: dict) -> dict | None:
    """Find the single best-matching row for a station spec."""
    rte = spec["rte"]
    desc_sub = spec["desc_contains"].upper()

    mask = (nev_df["RTE"] == rte)
    candidates = nev_df[mask]
    if candidates.empty:
        return None

    for _, row in candidates.iterrows():
        desc = str(row.get("DESCRIPTION", "")).strip().upper()
        if desc_sub in desc:
            back_aadt = _safe_int(row.get("BACK_AADT"))
            ahead_aadt = _safe_int(row.get("AHEAD_AADT"))
            vals = [v for v in (back_aadt, ahead_aadt) if v is not None]
            if not vals:
                return None
            observed_aadt = max(vals)
            return {
                "description": str(row.get("DESCRIPTION", "")).strip(),
                "back_aadt": back_aadt,
                "ahead_aadt": ahead_aadt,
                "observed_aadt": observed_aadt,
                "postmile": row.get("PM"),
            }
    return None


def find_best_model_link(conn, link_volumes: pd.DataFrame,
                         osm_names: list[str],
                         bbox: tuple[float, float, float, float]) -> dict | None:
    """
    Find the highest-volume model link within `bbox` whose name matches
    any of `osm_names`.  Returns dict with link_id, name, pce_tot, lon, lat.
    """
    xmin, ymin, xmax, ymax = bbox
    name_clauses = " OR ".join([f"name = ?" for _ in osm_names])
    sql = f"""
        SELECT link_id, name, link_type,
               X(Centroid(geometry)) AS cx, Y(Centroid(geometry)) AS cy
        FROM links
        WHERE ({name_clauses})
          AND X(Centroid(geometry)) BETWEEN ? AND ?
          AND Y(Centroid(geometry)) BETWEEN ? AND ?
    """
    params = list(osm_names) + [xmin, xmax, ymin, ymax]
    rows = conn.execute(sql, params).fetchall()
    if not rows:
        return None

    best = None
    best_vol = -1
    for lid, name, lt, cx, cy in rows:
        vol_match = link_volumes.loc[link_volumes["link_id"] == lid, "PCE_tot"]
        vol = float(vol_match.iloc[0]) if len(vol_match) else 0.0
        if vol > best_vol:
            best_vol = vol
            best = {"link_id": lid, "name": name, "link_type": lt,
                    "pce_tot": round(vol), "lon": cx, "lat": cy}
    return best


# ── main ─────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("STEP 3: OBSERVED-COUNT VALIDATION")
    print("=" * 60)

    # 1. Load Caltrans workbook
    print("\n[1] Loading Caltrans 2023 AADT workbook...")
    wb_bytes = download_workbook()
    nev_df = extract_nev_rows(wb_bytes)
    print(f"  {len(nev_df)} Nevada County rows extracted")

    # 2. Load model outputs
    print("\n[2] Loading canonical run outputs...")
    lv_path = OUT_DIR / "link_volumes.csv"
    geojson_path = OUT_DIR / "top_loaded_links.geojson"
    evidence_path = OUT_DIR / "evidence_packet.json"

    if not lv_path.exists():
        print(f"  ERROR: {lv_path} not found. Run step2_assign.py first.")
        sys.exit(1)

    link_volumes = pd.read_csv(lv_path)
    with open(evidence_path) as f:
        evidence = json.load(f)
    with open(geojson_path) as f:
        top_links_geojson = json.load(f)

    print(f"  link_volumes: {len(link_volumes)} rows, "
          f"{(link_volumes['PCE_tot'] > 0).sum()} with volume > 0")
    print(f"  evidence run_id: {evidence.get('run_id', 'unknown')}")

    # 3. Open network DB for spatial matching
    print("\n[3] Matching count stations to model links...")
    conn = sqlite3.connect(str(DB_PATH))
    conn.enable_load_extension(True)
    conn.load_extension("/home/linuxbrew/.linuxbrew/lib/mod_spatialite")

    results = []
    for station in STATIONS:
        print(f"\n  --- {station['label']} ---")

        # Caltrans observed
        ct = match_caltrans_row(nev_df, station["caltrans"])
        if ct is None:
            print(f"    WARNING: no Caltrans row found")
            results.append({
                "station_id": station["station_id"],
                "label": station["label"],
                "match_status": "caltrans_miss",
            })
            continue
        print(f"    Caltrans: {ct['description']}, "
              f"AADT={ct['observed_aadt']:,}, PM={ct['postmile']}")

        # Model link match
        ml = find_best_model_link(conn, link_volumes,
                                  station["osm_names"], station["bbox"])
        if ml is None:
            print(f"    WARNING: no model link found in bbox")
            results.append({
                "station_id": station["station_id"],
                "label": station["label"],
                "match_status": "model_miss",
                "observed_aadt": ct["observed_aadt"],
                "caltrans_description": ct["description"],
            })
            continue

        # The model volume (PCE_tot) is a daily-equivalent from a single
        # all-or-nothing/BFW assignment of LODES-based trips.
        # It is NOT an AADT; it is a screening-grade loaded volume.
        modeled_daily = ml["pce_tot"]

        # Compute comparison metrics
        if ct["observed_aadt"] > 0:
            abs_diff = abs(modeled_daily - ct["observed_aadt"])
            pct_error = 100.0 * abs_diff / ct["observed_aadt"]
            ratio = modeled_daily / ct["observed_aadt"]
        else:
            abs_diff = modeled_daily
            pct_error = None
            ratio = None

        print(f"    Model:    link_id={ml['link_id']}, name={ml['name']!r}, "
              f"PCE_tot={modeled_daily:,}")
        print(f"    Observed: {ct['observed_aadt']:,}  |  "
              f"Modeled: {modeled_daily:,}  |  "
              f"APE: {pct_error:.1f}%" if pct_error is not None else "APE: N/A")

        results.append({
            "station_id": station["station_id"],
            "label": station["label"],
            "match_status": "matched",
            "caltrans_description": ct["description"],
            "caltrans_postmile": ct["postmile"],
            "observed_aadt": ct["observed_aadt"],
            "model_link_id": ml["link_id"],
            "model_link_name": ml["name"],
            "model_link_type": ml["link_type"],
            "model_lon": round(ml["lon"], 5),
            "model_lat": round(ml["lat"], 5),
            "modeled_daily_pce": modeled_daily,
            "absolute_difference": round(abs_diff),
            "absolute_percent_error": round(pct_error, 2) if pct_error else None,
            "volume_ratio_model_obs": round(ratio, 4) if ratio else None,
        })

    conn.close()

    # 4. Compute summary metrics
    print("\n" + "=" * 60)
    print("[4] Computing summary metrics...")

    results_df = pd.DataFrame(results)
    matched = results_df[results_df["match_status"] == "matched"].copy()

    if matched.empty:
        print("  ERROR: no stations matched. Cannot compute metrics.")
        sys.exit(1)

    apes = matched["absolute_percent_error"].dropna()
    median_ape = float(apes.median())
    mean_ape = float(apes.mean())
    max_ape = float(apes.max())
    min_ape = float(apes.min())

    # Facility ranking: rank both observed and modeled, see if ordering
    # is preserved (Spearman rank correlation)
    matched_sorted = matched.sort_values("observed_aadt", ascending=False)
    obs_rank = matched_sorted["observed_aadt"].rank(ascending=False, method="min")
    mod_rank = matched_sorted["modeled_daily_pce"].rank(ascending=False, method="min")

    # Spearman rho (manual, small n)
    n = len(obs_rank)
    if n > 1:
        d_sq = ((obs_rank.values - mod_rank.values) ** 2).sum()
        spearman_rho = 1.0 - (6.0 * d_sq) / (n * (n**2 - 1))
    else:
        spearman_rho = None

    ranking_table = []
    for _, row in matched_sorted.iterrows():
        ranking_table.append({
            "station": row["label"],
            "observed_aadt": int(row["observed_aadt"]),
            "modeled_daily_pce": int(row["modeled_daily_pce"]),
            "obs_rank": int(obs_rank.loc[row.name]),
            "mod_rank": int(mod_rank.loc[row.name]),
        })

    print(f"\n  Stations matched: {len(matched)} / {len(STATIONS)}")
    print(f"  Median APE: {median_ape:.1f}%")
    print(f"  Mean APE:   {mean_ape:.1f}%")
    print(f"  Min APE:    {min_ape:.1f}%")
    print(f"  Max APE:    {max_ape:.1f}%")
    if spearman_rho is not None:
        print(f"  Spearman ρ (facility ranking): {spearman_rho:.3f}")

    print("\n  Facility Ranking Comparison:")
    print(f"  {'Station':<30s}  {'Obs AADT':>10s}  {'Mod PCE':>10s}  "
          f"{'Obs Rank':>8s}  {'Mod Rank':>8s}")
    for r in ranking_table:
        print(f"  {r['station']:<30s}  {r['observed_aadt']:>10,d}  "
              f"{r['modeled_daily_pce']:>10,d}  "
              f"{r['obs_rank']:>8d}  {r['mod_rank']:>8d}")

    # 5. Write outputs
    print(f"\n[5] Writing outputs...")

    csv_path = OUT_DIR / "validation_results.csv"
    results_df.to_csv(csv_path, index=False)
    print(f"  {csv_path}")

    summary = {
        "validation_type": "screening_assignment_vs_caltrans_aadt",
        "caltrans_year": 2023,
        "model_run_id": evidence.get("run_id", "unknown"),
        "model_engine": evidence.get("engine", "unknown"),
        "model_caveats": evidence.get("caveats", []),
        "stations_total": len(STATIONS),
        "stations_matched": int(len(matched)),
        "stations_missed": int(len(results_df) - len(matched)),
        "metrics": {
            "median_absolute_percent_error": round(median_ape, 2),
            "mean_absolute_percent_error": round(mean_ape, 2),
            "min_absolute_percent_error": round(min_ape, 2),
            "max_absolute_percent_error": round(max_ape, 2),
            "spearman_rho_facility_ranking": (
                round(spearman_rho, 4) if spearman_rho is not None else None
            ),
        },
        "facility_ranking": ranking_table,
        "interpretation": (
            "This is a screening-grade comparison between an uncalibrated "
            "LODES-based static assignment (BFW on OSM network with default "
            "speeds/capacities) and Caltrans 2023 observed AADT. High APE is "
            "expected: the model uses a synthetic OD matrix derived from "
            "workplace-area commute flows scaled by a flat factor, not a "
            "calibrated behavioral model. The facility-ranking correlation "
            "indicates whether the model correctly identifies relative demand "
            "magnitude across corridors, which is the useful screening signal."
        ),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    json_path = OUT_DIR / "validation_summary.json"
    with open(json_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"  {json_path}")

    print("\n" + "=" * 60)
    print("✅ VALIDATION COMPLETE")
    print("=" * 60)

    return 0


if __name__ == "__main__":
    sys.exit(main())
