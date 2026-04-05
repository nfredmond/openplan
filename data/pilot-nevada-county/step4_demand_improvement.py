#!/usr/bin/env python3
"""
Step 4: Improved demand matrix — NCHRP 716 / NHTS trip-rate expansion.

The LODES-only OD matrix (step 2) captures only home-based-work commute trips
at roughly 16,600 base trips (66k after a flat 4× factor).  Real daily person-
trips in Nevada County are on the order of 400k–450k because non-work purposes
(shopping, school, medical, social/recreational, non-home-based) dominate rural
travel.

This script:
  1. Keeps the LODES commute OD as the HBW layer (no flat factor).
  2. Generates HBO (home-based-other) and NHB (non-home-based) trip ends
     using NCHRP 716 / NHTS-rural trip rates applied to zone population
     and employment by sector.
  3. Distributes HBO and NHB trips via a doubly-constrained gravity model
     with travel-time impedance from the step-2 skim.
  4. Adds a simple external-trip cordon layer for SR 20, SR 49, and SR 174
     to represent through- and external-internal traffic.
  5. Sums all layers into a combined OD matrix.
  6. Runs a fresh BFW assignment on the existing network.
  7. Re-runs validation against Caltrans AADT.

Outputs go to run_output_v2/.
"""

import json
import os
import sqlite3
import sys
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import requests
import openpyxl
from scipy.spatial.distance import cdist

os.environ["SPATIALITE_LIBRARY_PATH"] = "/home/linuxbrew/.linuxbrew/lib/mod_spatialite"

from aequilibrae import Project
from aequilibrae.matrix import AequilibraeMatrix
from aequilibrae.paths import TrafficAssignment, TrafficClass, NetworkSkimming

# ── paths ────────────────────────────────────────────────────────────
DATA_DIR = Path(__file__).resolve().parent
PKG_DIR = DATA_DIR / "package"
OUT_DIR = DATA_DIR / "run_output_v2"
PROJ_DIR = DATA_DIR / "aeq_project"
OLD_OUT = DATA_DIR / "run_output"
DB_PATH = PROJ_DIR / "project_database.sqlite"

OUT_DIR.mkdir(exist_ok=True)

# ── constants ────────────────────────────────────────────────────────
# NCHRP 716 / NHTS rural trip-rate parameters (daily person-trips)
# These are consensus rates for small-urban / rural areas.
#
# Productions (trips produced per unit):
#   HBW:  already observed in LODES, use directly
#   HBO:  ~2.2 trips/person/day  (shopping + social/rec + personal business + school)
#   NHB:  ~0.9 trips/person/day
#
# Attractions (trips attracted per unit):
#   HBO:  retail emp × 12  +  service emp × 5  +  population × 0.5
#   NHB:  total emp × 2.5
#
# These are within the range documented in NCHRP 716 Table 4.4 and
# NHTS 2017 rural-area trip-rate summaries.

HBO_PROD_RATE = 2.2     # per resident
NHB_PROD_RATE = 0.9     # per resident

# Attraction weights
HBO_ATTR_RETAIL_RATE = 12.0
HBO_ATTR_SERVICE_RATE = 5.0   # health + accommodation + education
HBO_ATTR_POP_RATE = 0.5
NHB_ATTR_EMP_RATE = 2.5

# Gravity model impedance: f(t) = t^(-gamma)
GAMMA_HBO = 1.5
GAMMA_NHB = 1.2

# External trip parameters
# Based on cordon-count ratios typical for rural Sierra foothill counties:
# external-internal + through traffic is roughly 40-60% of total link volume
# on state highways.  We model this as fixed daily trip injections at gateway
# zones connecting to the nearest internal centroid.
#
# Gateway estimates derived from Caltrans AADT at county-line stations:
EXTERNAL_GATEWAYS = [
    # (label, nearest_zone_id, daily_trips_entering, daily_trips_exiting)
    # SR 49 south (Placer/Nevada CL): AADT ~27,000 → half entering
    ("SR49_south", 2, 13500, 13500),
    # SR 20 west (Yuba/Nevada CL): AADT ~7,900
    ("SR20_west", 3, 3950, 3950),
    # SR 174 south (Placer/Nevada CL): AADT ~5,600
    ("SR174_south", 22, 2800, 2800),
    # SR 20 east (toward I-80): AADT ~3,200
    ("SR20_east", 16, 1600, 1600),
    # SR 49 north (toward Downieville): AADT ~2,450
    ("SR49_north", 26, 1225, 1225),
    # I-80 corridor (tracts 6,9,10,12,13,17 are in the Truckee/Donner area
    # but these are far east; I-80 through-traffic is largely self-contained
    # in that sub-area).  We add a modest injection to tract 14 (large rural).
    ("I80_east", 14, 4000, 4000),
]


def load_zone_attrs() -> pd.DataFrame:
    """Load zone_attributes.csv, indexed by zone_id."""
    za = pd.read_csv(PKG_DIR / "zone_attributes.csv")
    za = za.set_index("zone_id").sort_index()
    return za


def load_lodes_od() -> np.ndarray:
    """Load the raw LODES OD (before any expansion factor) as an n×n array."""
    od = pd.read_csv(PKG_DIR / "od_trip_matrix.csv", index_col=0)
    # The saved matrix has the 4× factor baked in; divide it out
    return od.values / 4.0


def load_skim_matrix() -> np.ndarray:
    """Load the travel-time skim from the step-2 run."""
    from openmatrix import open_file
    omx_path = str(OLD_OUT / "travel_time_skims.omx")
    f = open_file(omx_path, "r")
    # The skim matrix name depends on what step2 used
    names = list(f.list_matrices())
    mat = np.array(f[names[0]])
    f.close()
    return mat


def gravity_distribute(productions: np.ndarray,
                       attractions: np.ndarray,
                       impedance: np.ndarray,
                       gamma: float,
                       max_iter: int = 50,
                       tolerance: float = 0.01) -> np.ndarray:
    """
    Doubly-constrained gravity model.
    
    Parameters
    ----------
    productions : 1-D array of trip productions per zone
    attractions : 1-D array of trip attractions per zone
    impedance   : n×n travel-time matrix
    gamma       : power-function exponent
    max_iter    : balancing iterations
    tolerance   : convergence threshold (max relative error)
    
    Returns
    -------
    n×n trip matrix
    """
    n = len(productions)
    
    # Friction factors: f(t) = t^(-gamma), with zero/inf protection
    with np.errstate(divide="ignore", invalid="ignore"):
        friction = np.where(
            (impedance > 0) & np.isfinite(impedance),
            impedance ** (-gamma),
            0.0,
        )
    # Zero out diagonal (no intra-zonal for this purpose)
    np.fill_diagonal(friction, 0.0)
    
    # Balance total attractions to total productions
    total_p = productions.sum()
    total_a = attractions.sum()
    if total_a > 0:
        attractions = attractions * (total_p / total_a)
    
    # Iterative balancing (Furness method)
    a_factor = np.ones(n)
    b_factor = np.ones(n)
    
    for iteration in range(max_iter):
        # Row (production) balancing
        seed = friction * a_factor[np.newaxis, :] * b_factor[:, np.newaxis]
        row_sums = seed.sum(axis=1)
        row_sums = np.where(row_sums > 0, row_sums, 1.0)
        b_factor = productions / row_sums
        
        # Column (attraction) balancing
        seed = friction * a_factor[np.newaxis, :] * b_factor[:, np.newaxis]
        col_sums = seed.sum(axis=0)
        col_sums = np.where(col_sums > 0, col_sums, 1.0)
        a_factor = attractions / col_sums
        
        # Check convergence
        trips = friction * a_factor[np.newaxis, :] * b_factor[:, np.newaxis]
        row_err = np.abs(trips.sum(axis=1) - productions)
        max_row_err = row_err.max() / max(productions.max(), 1)
        if max_row_err < tolerance:
            break
    
    result = friction * a_factor[np.newaxis, :] * b_factor[:, np.newaxis]
    return result


def build_external_matrix(n_zones: int, zone_ids: list[int]) -> np.ndarray:
    """
    Build an external trip matrix.
    
    External trips are injected as additional OD flows between gateway zones
    and all internal zones, weighted by zone employment (for entering trips)
    and zone population (for exiting trips).
    """
    za = load_zone_attrs()
    ext = np.zeros((n_zones, n_zones))
    
    for label, gateway_zid, daily_in, daily_out in EXTERNAL_GATEWAYS:
        gi = zone_ids.index(gateway_zid)
        
        # Entering trips: distribute from gateway to all zones by employment
        emp = za["total_jobs"].values.astype(float)
        emp_total = emp.sum()
        if emp_total > 0:
            shares_in = emp / emp_total
            ext[gi, :] += daily_in * shares_in
        
        # Exiting trips: distribute from all zones to gateway by population
        pop = za["est_population"].values.astype(float)
        pop_total = pop.sum()
        if pop_total > 0:
            shares_out = pop / pop_total
            ext[:, gi] += daily_out * shares_out
    
    return ext


def run_assignment(od_array: np.ndarray, centroids: list[int],
                   cost_field: str = "distance") -> pd.DataFrame:
    """
    Run BFW assignment on the existing network with the given OD matrix.
    Returns the link-results DataFrame.
    """
    project = Project()
    project.open(str(PROJ_DIR))
    project.network.build_graphs(modes=["c"])
    graph = project.network.graphs["c"]
    
    # Use same cost field logic as step2
    if "travel_time" in graph.graph.columns:
        cost_field = "travel_time"
    else:
        cost_field = "distance"
    
    graph.set_graph(cost_field)
    graph.prepare_graph(np.array(centroids))
    graph.set_blocked_centroid_flows(True)
    graph.set_skimming([cost_field])
    
    n_zones = len(centroids)
    
    # Create demand matrix (scaled to 10% peak hour for assignment)
    demand_path = str(OUT_DIR / "demand_v2.omx")
    demand_mat = AequilibraeMatrix()
    demand_mat.create_empty(
        file_name=demand_path, zones=n_zones,
        matrix_names=["demand"], memory_only=False,
    )
    demand_mat.index = np.array(centroids)
    # Scale to hourly peak (10% of daily) to avoid insane daily v/c ratios
    # since network capacity is hourly.
    peak_hour_factor = 0.10
    demand_mat.matrix["demand"][:, :] = od_array * peak_hour_factor
    demand_mat.computational_view(["demand"])
    
    # Assignment
    assig = TrafficAssignment()
    tc = TrafficClass(name="car", graph=graph, matrix=demand_mat)
    tc.set_pce(1.0)
    assig.add_class(tc)
    assig.set_vdf("BPR")
    assig.set_vdf_parameters({"alpha": 0.15, "beta": 4.0})
    assig.set_capacity_field("capacity")
    assig.set_time_field(cost_field)
    assig.max_iter = 50
    assig.rgap_target = 0.01
    assig.set_algorithm("bfw")
    
    assig.execute()
    
    rgap = getattr(assig.assignment, "rgap", float("nan"))
    iters = getattr(assig.assignment, "iteration", 50)
    print(f"  Assignment: gap={rgap:.6f}, iterations={iters}")
    
    results = assig.results()
    if hasattr(results, "get_load_results"):
        lr = results.get_load_results()
    else:
        # AequilibraE >= 1.1 returns DataFrame directly
        lr = results
    # Ensure link_id is a column, not just the index
    if "link_id" not in lr.columns and lr.index.name == "link_id":
        lr = lr.reset_index()
    elif "link_id" not in lr.columns:
        lr = lr.reset_index().rename(columns={"index": "link_id"})
    
    # Scale flows back up to daily (x10)
    if "PCE_tot" in lr.columns:
        lr["PCE_tot"] = lr["PCE_tot"] / peak_hour_factor
    
    # Also run skims on the congested network
    skimming = NetworkSkimming(graph)
    skimming.execute()
    skim_mat = skimming.results.skims
    skim_mat.export(str(OUT_DIR / "travel_time_skims_v2.omx"))
    
    project.close()
    return lr, rgap, iters, cost_field


def run_validation(link_volumes: pd.DataFrame, evidence: dict) -> dict:
    """
    Re-run the step3 validation logic against the new link volumes.
    Returns the summary dict.
    """
    # Import step3 logic inline to avoid module path issues
    CALTRANS_URL = (
        "https://dot.ca.gov/-/media/dot-media/programs/traffic-operations/"
        "documents/census/2023/2023-traffic-volumes-ca-a11y.xlsx"
    )
    CACHE_PATH = DATA_DIR / "caltrans_2023_aadt.xlsx"
    
    # Load cached workbook
    wb_bytes = CACHE_PATH.read_bytes()
    wb = openpyxl.load_workbook(BytesIO(wb_bytes), read_only=True, data_only=True)
    ws = wb["2023 AADT DATA"]
    header = [str(c.value).strip() if c.value else f"col_{i}"
              for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = dict(zip(header, row))
        if str(vals.get("CNTY", "")).strip().upper() == "NEV":
            records.append(vals)
    wb.close()
    nev_df = pd.DataFrame(records)
    nev_df["RTE"] = nev_df["RTE"].astype(str).str.strip().str.zfill(3)
    
    # Station definitions (same as step3)
    STATIONS = [
        {
            "station_id": "CT_SR20_JCT49",
            "label": "SR 20 at Jct Rte 49",
            "caltrans": {"rte": "020", "cnty": "NEV", "desc_contains": "JCT. RTE. 49"},
            "osm_names": ["Golden Center Freeway", "Golden Center Expressway",
                          "Grass Valley Highway"],
            "bbox": (-121.06, 39.215, -121.04, 39.235),
        },
        {
            "station_id": "CT_SR20_BRUNSWICK",
            "label": "SR 20 at Brunswick Rd",
            "caltrans": {"rte": "020", "cnty": "NEV", "desc_contains": "BRUNSWICK ROAD"},
            "osm_names": ["Grass Valley Highway", "Brunswick Road"],
            "bbox": (-121.05, 39.230, -121.03, 39.245),
        },
        {
            "station_id": "CT_SR20_PENNVALLEY",
            "label": "SR 20 at Penn Valley Dr",
            "caltrans": {"rte": "020", "cnty": "NEV", "desc_contains": "PENN VALLEY DRIVE"},
            "osm_names": ["Penn Valley Drive", "Grass Valley Highway"],
            "bbox": (-121.19, 39.190, -121.16, 39.210),
        },
        {
            "station_id": "CT_SR49_SOUTHGV",
            "label": "SR 49 at South Grass Valley",
            "caltrans": {"rte": "049", "cnty": "NEV", "desc_contains": "SOUTH GRASS VALLEY"},
            "osm_names": ["State Highway 49", "Alta Sierra Drive",
                          "Golden Center Freeway"],
            "bbox": (-121.08, 39.130, -121.04, 39.170),
        },
        {
            "station_id": "CT_SR174_BRUNSWICK",
            "label": "SR 174 at Brunswick Rd",
            "caltrans": {"rte": "174", "cnty": "NEV", "desc_contains": "BRUNSWICK ROAD"},
            "osm_names": ["Colfax Highway", "Brunswick Road"],
            "bbox": (-121.06, 39.200, -121.03, 39.225),
        },
    ]
    
    def _safe_int(v):
        if v is None: return None
        s = str(v).strip()
        if s in ("", " "): return None
        try: return int(float(s))
        except: return None
    
    conn = sqlite3.connect(str(DB_PATH))
    conn.enable_load_extension(True)
    conn.load_extension("/home/linuxbrew/.linuxbrew/lib/mod_spatialite")
    
    results = []
    for station in STATIONS:
        # Match Caltrans
        rte = station["caltrans"]["rte"]
        desc_sub = station["caltrans"]["desc_contains"].upper()
        mask = nev_df["RTE"] == rte
        ct_match = None
        for _, row in nev_df[mask].iterrows():
            desc = str(row.get("DESCRIPTION", "")).strip().upper()
            if desc_sub in desc:
                back = _safe_int(row.get("BACK_AADT"))
                ahead = _safe_int(row.get("AHEAD_AADT"))
                vals = [v for v in (back, ahead) if v is not None]
                if vals:
                    ct_match = {
                        "description": str(row.get("DESCRIPTION", "")).strip(),
                        "observed_aadt": max(vals),
                        "postmile": row.get("PM"),
                    }
                break
        
        if ct_match is None:
            results.append({"station_id": station["station_id"],
                          "label": station["label"], "match_status": "caltrans_miss"})
            continue
        
        # Match model link
        xmin, ymin, xmax, ymax = station["bbox"]
        name_clauses = " OR ".join(["name = ?" for _ in station["osm_names"]])
        sql = f"""SELECT link_id, name, link_type,
                         X(Centroid(geometry)) AS cx, Y(Centroid(geometry)) AS cy
                  FROM links
                  WHERE ({name_clauses})
                    AND X(Centroid(geometry)) BETWEEN ? AND ?
                    AND Y(Centroid(geometry)) BETWEEN ? AND ?"""
        params = list(station["osm_names"]) + [xmin, xmax, ymin, ymax]
        rows = conn.execute(sql, params).fetchall()
        
        best_link = None
        best_vol = -1
        for lid, name, lt, cx, cy in rows:
            vol_match = link_volumes.loc[link_volumes["link_id"] == lid, "PCE_tot"]
            vol = float(vol_match.iloc[0]) if len(vol_match) else 0.0
            if vol > best_vol:
                best_vol = vol
                best_link = {"link_id": lid, "name": name, "link_type": lt,
                            "pce_tot": round(vol), "lon": cx, "lat": cy}
        
        if best_link is None:
            results.append({"station_id": station["station_id"],
                          "label": station["label"], "match_status": "model_miss",
                          "observed_aadt": ct_match["observed_aadt"]})
            continue
        
        modeled = best_link["pce_tot"]
        obs = ct_match["observed_aadt"]
        abs_diff = abs(modeled - obs)
        pct_err = 100.0 * abs_diff / obs if obs > 0 else None
        ratio = modeled / obs if obs > 0 else None
        
        results.append({
            "station_id": station["station_id"],
            "label": station["label"],
            "match_status": "matched",
            "caltrans_description": ct_match["description"],
            "caltrans_postmile": ct_match["postmile"],
            "observed_aadt": obs,
            "model_link_id": best_link["link_id"],
            "model_link_name": best_link["name"],
            "model_link_type": best_link["link_type"],
            "model_lon": round(best_link["lon"], 5),
            "model_lat": round(best_link["lat"], 5),
            "modeled_daily_pce": modeled,
            "absolute_difference": round(abs_diff),
            "absolute_percent_error": round(pct_err, 2) if pct_err else None,
            "volume_ratio_model_obs": round(ratio, 4) if ratio else None,
        })
    
    conn.close()
    
    results_df = pd.DataFrame(results)
    matched = results_df[results_df["match_status"] == "matched"].copy()
    
    apes = matched["absolute_percent_error"].dropna()
    median_ape = float(apes.median())
    mean_ape = float(apes.mean())
    
    obs_rank = matched["observed_aadt"].rank(ascending=False, method="min")
    mod_rank = matched["modeled_daily_pce"].rank(ascending=False, method="min")
    n = len(obs_rank)
    if n > 1:
        d_sq = ((obs_rank.values - mod_rank.values) ** 2).sum()
        spearman_rho = 1.0 - (6.0 * d_sq) / (n * (n**2 - 1))
    else:
        spearman_rho = None
    
    ranking_table = []
    for idx in matched.sort_values("observed_aadt", ascending=False).index:
        row = matched.loc[idx]
        ranking_table.append({
            "station": row["label"],
            "observed_aadt": int(row["observed_aadt"]),
            "modeled_daily_pce": int(row["modeled_daily_pce"]),
            "obs_rank": int(obs_rank.loc[idx]),
            "mod_rank": int(mod_rank.loc[idx]),
        })
    
    return results_df, {
        "median_ape": median_ape,
        "mean_ape": mean_ape,
        "min_ape": float(apes.min()),
        "max_ape": float(apes.max()),
        "spearman_rho": spearman_rho,
        "ranking_table": ranking_table,
        "stations_matched": int(len(matched)),
    }


# ═════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("STEP 4: DEMAND IMPROVEMENT — NHTS TRIP-RATE + EXTERNAL TRIPS")
    print("=" * 65)
    
    # ── 1. Load inputs ───────────────────────────────────────────────
    print("\n[1] Loading inputs...")
    za = load_zone_attrs()
    zone_ids = sorted(za.index.tolist())
    n_zones = len(zone_ids)
    print(f"  {n_zones} zones")
    print(f"  Total population: {za['est_population'].sum():,}")
    print(f"  Total employment: {za['total_jobs'].sum():,}")
    
    lodes_od = load_lodes_od()
    print(f"  LODES base commute trips: {lodes_od.sum():,.0f}")
    
    skim = load_skim_matrix()
    print(f"  Skim matrix: {skim.shape}, "
          f"reachable pairs: {((skim > 0) & np.isfinite(skim)).sum()}")
    
    # ── 2. Trip generation ───────────────────────────────────────────
    print("\n[2] Trip generation (NCHRP 716 / NHTS rural rates)...")
    
    pop = za["est_population"].values.astype(float)
    total_emp = za["total_jobs"].values.astype(float)
    retail_emp = za["retail_jobs"].values.astype(float)
    service_emp = (za["health_jobs"].values + za["accommodation_jobs"].values +
                   za["education_jobs"].values).astype(float)
    
    # HBW: use LODES directly (no expansion factor)
    hbw_trips = lodes_od.sum()
    print(f"  HBW (LODES): {hbw_trips:,.0f} trips")
    
    # HBO productions and attractions
    hbo_prod = pop * HBO_PROD_RATE
    hbo_attr = (retail_emp * HBO_ATTR_RETAIL_RATE +
                service_emp * HBO_ATTR_SERVICE_RATE +
                pop * HBO_ATTR_POP_RATE)
    print(f"  HBO: {hbo_prod.sum():,.0f} productions, {hbo_attr.sum():,.0f} attractions")
    
    # NHB productions and attractions
    nhb_prod = pop * NHB_PROD_RATE
    nhb_attr = total_emp * NHB_ATTR_EMP_RATE
    print(f"  NHB: {nhb_prod.sum():,.0f} productions, {nhb_attr.sum():,.0f} attractions")
    
    # ── 3. Trip distribution (gravity model) ─────────────────────────
    print("\n[3] Trip distribution (doubly-constrained gravity model)...")
    
    hbo_od = gravity_distribute(hbo_prod, hbo_attr, skim, GAMMA_HBO)
    nhb_od = gravity_distribute(nhb_prod, nhb_attr, skim, GAMMA_NHB)
    
    print(f"  HBO distributed: {hbo_od.sum():,.0f} trips")
    print(f"  NHB distributed: {nhb_od.sum():,.0f} trips")
    
    # ── 4. External trips ────────────────────────────────────────────
    print("\n[4] Building external trip layer...")
    ext_od = build_external_matrix(n_zones, zone_ids)
    print(f"  External trips: {ext_od.sum():,.0f}")
    
    for label, gzid, d_in, d_out in EXTERNAL_GATEWAYS:
        gi = zone_ids.index(gzid)
        entering = ext_od[gi, :].sum()
        exiting = ext_od[:, gi].sum()
        print(f"    {label:15s}: entering={entering:,.0f}, exiting={exiting:,.0f}")
    
    # ── 5. Combine all layers ────────────────────────────────────────
    print("\n[5] Combining demand layers...")
    combined = lodes_od + hbo_od + nhb_od + ext_od
    
    # Zero out unreachable pairs
    unreachable = ~((skim > 0) & np.isfinite(skim))
    np.fill_diagonal(unreachable, True)
    combined[unreachable] = 0
    
    total_trips = combined.sum()
    print(f"  HBW:       {lodes_od[~unreachable].sum():>10,.0f}")
    print(f"  HBO:       {hbo_od[~unreachable].sum():>10,.0f}")
    print(f"  NHB:       {nhb_od[~unreachable].sum():>10,.0f}")
    print(f"  External:  {ext_od[~unreachable].sum():>10,.0f}")
    print(f"  ─────────────────────────")
    print(f"  Total:     {total_trips:>10,.0f}")
    
    # Save combined OD
    od_df = pd.DataFrame(combined, index=zone_ids, columns=zone_ids)
    od_df.index.name = "origin_zone"
    od_df.to_csv(OUT_DIR / "od_trip_matrix_v2.csv")
    
    # Save layer breakdown
    layer_summary = {
        "hbw_trips": round(float(lodes_od[~unreachable].sum())),
        "hbo_trips": round(float(hbo_od[~unreachable].sum())),
        "nhb_trips": round(float(nhb_od[~unreachable].sum())),
        "external_trips": round(float(ext_od[~unreachable].sum())),
        "total_trips": round(float(total_trips)),
        "trip_rates": {
            "hbo_prod_per_person": HBO_PROD_RATE,
            "nhb_prod_per_person": NHB_PROD_RATE,
            "gravity_gamma_hbo": GAMMA_HBO,
            "gravity_gamma_nhb": GAMMA_NHB,
        },
        "external_gateways": [
            {"label": g[0], "zone_id": g[1],
             "daily_in": g[2], "daily_out": g[3]}
            for g in EXTERNAL_GATEWAYS
        ],
    }
    with open(OUT_DIR / "demand_layers.json", "w") as f:
        json.dump(layer_summary, f, indent=2)
    
    # ── 6. Run assignment ────────────────────────────────────────────
    print("\n[6] Running BFW assignment on improved demand...")
    
    # Get centroid list (same as step2)
    conn = sqlite3.connect(str(DB_PATH))
    conn.enable_load_extension(True)
    conn.load_extension("/home/linuxbrew/.linuxbrew/lib/mod_spatialite")
    centroids = [r[0] for r in conn.execute(
        "SELECT node_id FROM nodes WHERE is_centroid=1 ORDER BY node_id"
    ).fetchall()]
    conn.close()
    
    print(f"  Centroids: {len(centroids)}")
    print(f"  Total demand: {combined.sum():,.0f} trips")
    
    lr, rgap, iters, cost_field = run_assignment(combined, centroids)
    
    # Save results
    lr.to_csv(OUT_DIR / "link_volumes.csv")
    loaded = lr[lr["PCE_tot"] > 0]
    print(f"  Links with volume: {len(loaded)}/{len(lr)}")
    
    if len(loaded) > 0:
        top10 = loaded.nlargest(10, "PCE_tot")
        print("\n  Top 10 loaded links:")
        for _, row in top10.iterrows():
            lid = int(row.get("link_id", row.name))
            print(f"    Link {lid}: {row['PCE_tot']:,.0f} PCE")
    
    # Build top-loaded GeoJSON (same as step2)
    conn = sqlite3.connect(str(DB_PATH))
    conn.enable_load_extension(True)
    conn.load_extension("/home/linuxbrew/.linuxbrew/lib/mod_spatialite")
    
    top20_ids = loaded.nlargest(20, "PCE_tot")
    features = []
    for _, row in top20_ids.iterrows():
        lid = int(row.get("link_id", row.name))
        geom_row = conn.execute(
            "SELECT name, link_type, AsGeoJSON(geometry) FROM links WHERE link_id=?",
            (lid,)
        ).fetchone()
        if geom_row:
            features.append({
                "type": "Feature",
                "properties": {
                    "link_id": lid,
                    "name": geom_row[0],
                    "link_type": geom_row[1],
                    "pce_tot": round(float(row["PCE_tot"])),
                },
                "geometry": json.loads(geom_row[2]),
            })
    conn.close()
    
    with open(OUT_DIR / "top_loaded_links.geojson", "w") as f:
        json.dump({"type": "FeatureCollection", "features": features}, f)
    
    # Evidence packet
    evidence = {
        "run_id": "nevada-county-pilot-v2-nhts-expansion",
        "engine": "AequilibraE 1.6.1",
        "network_source": "OpenStreetMap",
        "model_area": "Nevada County (-121.30,39.00 to -120.00,39.50)",
        "algorithm": "BFW",
        "vdf": "BPR (α=0.15, β=4.0)",
        "convergence": {
            "final_gap": float(rgap) if np.isfinite(rgap) else None,
            "iterations": int(iters),
        },
        "network": {"zones": len(centroids)},
        "demand": {
            "total_trips": round(float(total_trips)),
            "hbw_trips": round(float(lodes_od[~unreachable].sum())),
            "hbo_trips": round(float(hbo_od[~unreachable].sum())),
            "nhb_trips": round(float(nhb_od[~unreachable].sum())),
            "external_trips": round(float(ext_od[~unreachable].sum())),
            "source": "LODES 2021 HBW + NCHRP 716 HBO/NHB + external cordon",
        },
        "loaded_links": int(len(loaded)),
        "caveats": [
            "Uncalibrated (trip rates from national averages, not local survey)",
            "OSM default speeds/capacities",
            "Closed boundary with external injection approximation",
            "Screening-grade",
            "Gravity model with assumed gamma, not calibrated to observed TLD",
            "External trips allocated by simple employment/population shares",
        ],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    with open(OUT_DIR / "evidence_packet.json", "w") as f:
        json.dump(evidence, f, indent=2)
    
    # ── 7. Validation ────────────────────────────────────────────────
    print("\n[7] Running validation against Caltrans 2023 AADT...")
    
    val_df, val_metrics = run_validation(lr, evidence)
    val_df.to_csv(OUT_DIR / "validation_results.csv", index=False)
    
    summary = {
        "validation_type": "screening_assignment_vs_caltrans_aadt",
        "demand_version": "v2_nhts_expansion",
        "caltrans_year": 2023,
        "model_run_id": evidence["run_id"],
        "model_engine": evidence["engine"],
        "model_caveats": evidence["caveats"],
        "demand_summary": layer_summary,
        "stations_total": 5,
        "stations_matched": val_metrics["stations_matched"],
        "metrics": {
            "median_absolute_percent_error": round(val_metrics["median_ape"], 2),
            "mean_absolute_percent_error": round(val_metrics["mean_ape"], 2),
            "min_absolute_percent_error": round(val_metrics["min_ape"], 2),
            "max_absolute_percent_error": round(val_metrics["max_ape"], 2),
            "spearman_rho_facility_ranking": (
                round(val_metrics["spearman_rho"], 4)
                if val_metrics["spearman_rho"] is not None else None
            ),
        },
        "facility_ranking": val_metrics["ranking_table"],
        "comparison_to_v1": {
            "v1_median_ape": 97.78,
            "v1_spearman_rho": -0.9,
            "v2_median_ape": round(val_metrics["median_ape"], 2),
            "v2_spearman_rho": (
                round(val_metrics["spearman_rho"], 4)
                if val_metrics["spearman_rho"] is not None else None
            ),
        },
        "interpretation": (
            "The improved demand model adds NCHRP 716 / NHTS-based non-work "
            "trip generation (HBO, NHB) plus external-trip cordon injections "
            "to the LODES commute base. This materially increases total demand "
            "and should shift corridor-level volumes closer to observed AADT. "
            "However, the model remains uncalibrated: trip rates are national "
            "averages, the gravity model gamma is assumed, and external trip "
            "volumes are rough estimates from Caltrans county-line stations. "
            "APE improvement indicates the demand gap is closing; ranking "
            "improvement indicates the spatial distribution is improving."
        ),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    with open(OUT_DIR / "validation_summary.json", "w") as f:
        json.dump(summary, f, indent=2)
    
    # ── Report ───────────────────────────────────────────────────────
    print("\n" + "=" * 65)
    print("VALIDATION RESULTS (v2 vs v1)")
    print("=" * 65)
    print(f"\n  {'Station':<30s}  {'Obs AADT':>10s}  {'v1 PCE':>8s}  "
          f"{'v2 PCE':>8s}  {'v1 APE':>7s}  {'v2 APE':>7s}")
    
    v1_pce = {"SR 20 at Jct Rte 49": 22, "SR 20 at Brunswick Rd": 151,
              "SR 20 at Penn Valley Dr": 518, "SR 49 at South Grass Valley": 578,
              "SR 174 at Brunswick Rd": 643}
    
    for r in val_metrics["ranking_table"]:
        obs = r["observed_aadt"]
        v2 = r["modeled_daily_pce"]
        v1 = v1_pce.get(r["station"], 0)
        v1_ape = 100 * abs(v1 - obs) / obs if obs > 0 else 0
        v2_ape = 100 * abs(v2 - obs) / obs if obs > 0 else 0
        print(f"  {r['station']:<30s}  {obs:>10,d}  {v1:>8,d}  "
              f"{v2:>8,d}  {v1_ape:>6.1f}%  {v2_ape:>6.1f}%")
    
    print(f"\n  v1 Median APE: 97.78%  →  v2 Median APE: {val_metrics['median_ape']:.1f}%")
    rho = val_metrics['spearman_rho']
    rho_str = f"{rho:.3f}" if rho is not None else "N/A"
    print(f"  v1 Spearman ρ: -0.900  →  v2 Spearman ρ: {rho_str}")
    
    print(f"\n  Output directory: {OUT_DIR}")
    print("=" * 65)
    print("✅ STEP 4 COMPLETE")
    print("=" * 65)
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
